"""
generer_formulaire_excel.py
Convertit formulaire_notation.json en tableur Excel pour notation humaine.
Usage : python evaluation/generer_formulaire_excel.py
"""

import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ── Chemins ───────────────────────────────────────────────────────
INPUT  = Path("C:\\Users\\Samir\\Documents\\LRGP-PROJECT\\lrgp_llm\\rag\\evaluation\\results\\formulaire_notation_v4_with_rag.json")
OUTPUT = Path("C:\\Users\\Samir\\Documents\\LRGP-PROJECT\\lrgp_llm\\rag\\evaluation\\results\\formulaire_notation_v4_with_rag.xlsx")

# ── Couleurs ──────────────────────────────────────────────────────
NAVY   = "1F3864"
BLUE   = "2563A8"
TEAL   = "0D9488"
AMBER  = "FFF2CC"
GREEN  = "E2F0D9"
RED    = "FFE0E0"
GREY   = "F2F2F2"
WHITE  = "FFFFFF"
LIGHT  = "EEF3FA"

def border(style="thin", color="CCCCCC"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def hcell(ws, row, col, text, bg=NAVY, fg=WHITE,
          bold=True, size=10, wrap=False, align="center"):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(bold=bold, color=fg, size=size, name="Arial")
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center",
                             wrap_text=wrap)
    c.border    = border()
    return c

def dcell(ws, row, col, text="", bg=WHITE, bold=False,
          size=10, wrap=True, align="left", color="000000"):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(bold=bold, color=color, size=size, name="Arial")
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="top",
                             wrap_text=wrap)
    c.border    = border()
    return c

def main():
    with open(INPUT, encoding="utf-8") as f:
        fiches = json.load(f)

    wb = Workbook()

    # ══════════════════════════════════════════════════════════════
    # ONGLET 1 — GUIDE DE NOTATION
    # ══════════════════════════════════════════════════════════════
    ws_guide = wb.active
    ws_guide.title = "Guide de notation"
    ws_guide.column_dimensions["A"].width = 28
    ws_guide.column_dimensions["B"].width = 55
    ws_guide.column_dimensions["C"].width = 20

    # Titre
    ws_guide.merge_cells("A1:C1")
    c = ws_guide["A1"]
    c.value     = "GRILLE DE NOTATION — Assistant IA LRGP"
    c.font      = Font(bold=True, size=16, color=WHITE, name="Arial")
    c.fill      = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_guide.row_dimensions[1].height = 40

    ws_guide.merge_cells("A2:C2")
    c = ws_guide["A2"]
    c.value     = "Chaque critère est noté de 1 à 5 — 5 = parfait, 1 = incorrect"
    c.font      = Font(italic=True, size=11, color="444444", name="Arial")
    c.fill      = PatternFill("solid", fgColor=LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_guide.row_dimensions[2].height = 25

    # En-têtes
    for col, txt in enumerate(["Critère", "Description", "Barème"], 1):
        hcell(ws_guide, 3, col, txt, bg=BLUE)
    ws_guide.row_dimensions[3].height = 20

    criteres = [
        ("Exactitude factuelle",
         "Les valeurs numériques sont-elles correctes ?\n"
         "Les faits scientifiques sont-ils justes ?\n"
         "5 = toutes les valeurs exactes\n"
         "3 = quelques erreurs mineures\n"
         "1 = valeurs incorrectes ou inventées",
         "1 = Faux\n2 = Insuffisant\n3 = Partiel\n4 = Bon\n5 = Parfait"),

        ("Rigueur de la démarche",
         "Les étapes de calcul sont-elles correctes et complètes ?\n"
         "La démarche est-elle reproductible ?\n"
         "5 = toutes les étapes détaillées et correctes\n"
         "3 = étapes présentes mais incomplètes\n"
         "1 = démarche absente ou erronée",
         "1 = Absente\n2 = Insuffisant\n3 = Partielle\n4 = Bonne\n5 = Complète"),

        ("Pertinence physique",
         "Les unités sont-elles correctes à chaque étape ?\n"
         "Les ordres de grandeur sont-ils cohérents ?\n"
         "5 = unités et ordres de grandeur parfaits\n"
         "3 = quelques erreurs d'unités\n"
         "1 = erreurs d'unités majeures",
         "1 = Faux\n2 = Insuffisant\n3 = Partiel\n4 = Bon\n5 = Parfait"),

        ("Clarté pédagogique",
         "La réponse est-elle bien structurée et lisible ?\n"
         "Un étudiant M2 comprendrait-il ?\n"
         "5 = claire, structurée, exemplaire\n"
         "3 = compréhensible mais améliorable\n"
         "1 = confuse ou illisible",
         "1 = Confuse\n2 = Insuffisant\n3 = Acceptable\n4 = Claire\n5 = Exemplaire"),

        ("Citation des sources",
         "Les sources sont-elles citées ?\n"
         "Les citations sont-elles précises et correctes ?\n"
         "5 = toutes les sources citées correctement\n"
         "3 = sources partiellement citées\n"
         "1 = aucune source citée",
         "1 = Absente\n2 = Insuffisant\n3 = Partielle\n4 = Bonne\n5 = Complète"),
    ]

    for i, (nom, desc, bareme) in enumerate(criteres, 4):
        bg = GREY if i % 2 == 0 else WHITE
        dcell(ws_guide, i, 1, nom,    bg=bg, bold=True, size=10)
        dcell(ws_guide, i, 2, desc,   bg=bg, size=9,  wrap=True)
        dcell(ws_guide, i, 3, bareme, bg=bg, size=9,  wrap=True)
        ws_guide.row_dimensions[i].height = 70

    # Section calibration
    r = len(criteres) + 5
    ws_guide.merge_cells(f"A{r}:C{r}")
    c = ws_guide.cell(row=r, column=1,
                      value="PROTOCOLE DE CALIBRATION (30 min)")
    c.font      = Font(bold=True, size=12, color=WHITE, name="Arial")
    c.fill      = PatternFill("solid", fgColor=TEAL)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_guide.row_dimensions[r].height = 25

    etapes = [
        "1. Les deux annotateurs notent les 5 premières fiches ensemble à voix haute",
        "2. Discussion des désaccords pour aligner la compréhension des critères",
        "3. Notation indépendante des 36 fiches restantes sans se concerter",
        "4. Calcul du κ Cohen après notation complète (script compute_kappa.py)",
        "5. κ > 0.6 → accord satisfaisant | κ < 0.4 → recalibrer",
    ]
    for j, etape in enumerate(etapes, r+1):
        ws_guide.merge_cells(f"A{j}:C{j}")
        c = ws_guide.cell(row=j, column=1, value=etape)
        c.font      = Font(size=10, name="Arial")
        c.fill      = PatternFill("solid", fgColor=LIGHT)
        c.alignment = Alignment(horizontal="left", vertical="center",
                                 indent=1)
        c.border    = border()
        ws_guide.row_dimensions[j].height = 20

    # ══════════════════════════════════════════════════════════════
    # ONGLET 2 — ANNOTATEUR 1
    # ONGLET 3 — ANNOTATEUR 2
    # ══════════════════════════════════════════════════════════════
    CRITERES_COLS = [
        "exactitude_factuelle",
        "rigueur_demarche",
        "pertinence_physique",
        "clarte_pedagogique",
        "citation_sources",
    ]
    CRITERES_LABELS = [
        "Exactitude\nfactuelle",
        "Rigueur\ndémarche",
        "Pertinence\nphysique",
        "Clarté\npédagogique",
        "Citation\nsources",
    ]

    for annotateur in [1, 2]:
        ws = wb.create_sheet(title=f"Annotateur {annotateur}")

        # Largeurs colonnes
        largeurs = {
            "A": 8,   # ID
            "B": 10,  # Type
            "C": 12,  # Niveau
            "D": 45,  # Question
            "E": 60,  # Réponse modèle
            "F": 35,  # Référence
            "G": 14,  # Exactitude
            "H": 14,  # Rigueur
            "I": 14,  # Pertinence
            "J": 14,  # Clarté
            "K": 14,  # Sources
            "L": 12,  # Score global
            "M": 30,  # Commentaire
        }
        for col_letter, width in largeurs.items():
            ws.column_dimensions[col_letter].width = width

        # ── Titre ─────────────────────────────────────────────────
        ws.merge_cells("A1:M1")
        c = ws["A1"]
        c.value     = f"ÉVALUATION HUMAINE — ANNOTATEUR {annotateur} — Baseline RAG Qwen3.5:9b"
        c.font      = Font(bold=True, size=14, color=WHITE, name="Arial")
        c.fill      = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 35

        # Sous-titre
        ws.merge_cells("A2:M2")
        c = ws["A2"]
        c.value     = ("Remplir les colonnes G à L avec des notes de 1 à 5 "
                       "— Colonne L calculée automatiquement — "
                       "Ne pas consulter l'autre annotateur")
        c.font      = Font(italic=True, size=10, color="444444", name="Arial")
        c.fill      = PatternFill("solid", fgColor=AMBER)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 20

        # ── En-têtes colonnes ─────────────────────────────────────
        headers = [
            ("A", 3, "ID",         NAVY),
            ("B", 3, "Type",       NAVY),
            ("C", 3, "Niveau",     NAVY),
            ("D", 3, "Question",   NAVY),
            ("E", 3, "Réponse modèle", NAVY),
            ("F", 3, "Référence",  NAVY),
            ("G", 3, CRITERES_LABELS[0], TEAL),
            ("H", 3, CRITERES_LABELS[1], TEAL),
            ("I", 3, CRITERES_LABELS[2], TEAL),
            ("J", 3, CRITERES_LABELS[3], TEAL),
            ("K", 3, CRITERES_LABELS[4], TEAL),
            ("L", 3, "Score\nglobal",    "0D7377"),
            ("M", 3, "Commentaire",      BLUE),
        ]
        for col_l, row, txt, bg in headers:
            col_n = ord(col_l) - ord("A") + 1
            hcell(ws, row, col_n, txt, bg=bg, size=9, wrap=True)
        ws.row_dimensions[3].height = 35

        # ── Fixer les 3 premières lignes ──────────────────────────
        ws.freeze_panes = "A4"

        # ── Données ───────────────────────────────────────────────
        TYPE_COLORS = {
            "CALCUL":      "DEEAF1",
            "FACTUEL":     "E2F0D9",
            "COMPARAISON": "FFF2CC",
        }
        NIVEAU_COLORS = {
            "N1": "E2F0D9",
            "N2": "DEEAF1",
            "N3": "FFF2CC",
            "N4": "FFE0E0",
        }

        for i, fiche in enumerate(fiches):
            row = i + 4
            bg_row = GREY if i % 2 == 0 else WHITE
            annot_key = f"notation_annotateur_{annotateur}"
            notation   = fiche.get(annot_key, {})

            # Nettoyer la réponse modèle — supprimer markdown lourd
            reponse = fiche.get("reponse_modele", "")
            reponse = reponse.replace("**", "").replace("##", "")

            # ID
            dcell(ws, row, 1, fiche["id"], bg=bg_row,
                  bold=True, align="center")

            # Type
            bg_type = TYPE_COLORS.get(fiche.get("type",""), bg_row)
            dcell(ws, row, 2, fiche.get("type",""), bg=bg_type,
                  align="center", bold=True, size=9)

            # Niveau
            bg_niv = NIVEAU_COLORS.get(fiche.get("difficulté",""), bg_row)
            dcell(ws, row, 3, fiche.get("difficulté",""), bg=bg_niv,
                  align="center", bold=True, size=9)

            # Question
            dcell(ws, row, 4, fiche["question"], bg=bg_row, size=9)

            # Réponse modèle
            dcell(ws, row, 5, reponse[:2000], bg=bg_row, size=9)

            # Référence
            dcell(ws, row, 6, fiche.get("reference",""), bg=LIGHT, size=9)

            # Colonnes de notation G-K
            for j, cle in enumerate(CRITERES_COLS):
                val = notation.get(cle)
                col_n = 7 + j
                c = ws.cell(row=row, column=col_n, value=val)
                c.font      = Font(size=11, bold=True, name="Arial",
                                   color="1F3864")
                c.fill      = PatternFill("solid",
                                fgColor="FFFDE7" if val is None else "E8F5E9")
                c.alignment = Alignment(horizontal="center",
                                        vertical="center")
                c.border    = border(color="999999")

            # Score global — formule moyenne
            g_col = get_column_letter(7)
            k_col = get_column_letter(11)
            c = ws.cell(row=row, column=12,
                        value=f"=IFERROR(AVERAGE({g_col}{row}:{k_col}{row}),\"\")")
            c.font      = Font(size=11, bold=True, name="Arial",
                               color=WHITE)
            c.fill      = PatternFill("solid", fgColor=TEAL)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = border()

            # Commentaire
            commentaire = notation.get("commentaire", "")
            dcell(ws, row, 13, commentaire, bg=bg_row, size=9)

            ws.row_dimensions[row].height = 80

        # ── Ligne de synthèse ─────────────────────────────────────
        last = len(fiches) + 4
        ws.merge_cells(f"A{last}:F{last}")
        c = ws.cell(row=last, column=1, value="SCORE MOYEN GLOBAL")
        c.font      = Font(bold=True, size=11, color=WHITE, name="Arial")
        c.fill      = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="right", vertical="center")

        # Moyenne de chaque critère
        for j in range(5):
            col_n = 7 + j
            col_l = get_column_letter(col_n)
            c = ws.cell(row=last, column=col_n,
                        value=f"=IFERROR(AVERAGE({col_l}4:{col_l}{last-1}),\"\")")
            c.font      = Font(bold=True, size=11, name="Arial", color=WHITE)
            c.fill      = PatternFill("solid", fgColor=TEAL)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = border()
            c.number_format = "0.00"

        # Moyenne globale
        c = ws.cell(row=last, column=12,
                    value=f"=IFERROR(AVERAGE(L4:L{last-1}),\"\")")
        c.font      = Font(bold=True, size=12, name="Arial", color=WHITE)
        c.fill      = PatternFill("solid", fgColor="0D7377")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = border()
        c.number_format = "0.00"

        ws.row_dimensions[last].height = 25

    # ══════════════════════════════════════════════════════════════
    # ONGLET 4 — SYNTHÈSE
    # ══════════════════════════════════════════════════════════════
    ws_synth = wb.create_sheet(title="Synthèse")
    ws_synth.column_dimensions["A"].width = 30
    ws_synth.column_dimensions["B"].width = 20
    ws_synth.column_dimensions["C"].width = 20
    ws_synth.column_dimensions["D"].width = 20

    ws_synth.merge_cells("A1:D1")
    c = ws_synth["A1"]
    c.value     = "SYNTHÈSE — Accord inter-annotateurs"
    c.font      = Font(bold=True, size=14, color=WHITE, name="Arial")
    c.fill      = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_synth.row_dimensions[1].height = 35

    for col, txt in enumerate(["Critère","Annotateur 1","Annotateur 2","Écart"], 1):
        hcell(ws_synth, 2, col, txt, bg=BLUE)

    criteres_noms = [
        "Exactitude factuelle",
        "Rigueur démarche",
        "Pertinence physique",
        "Clarté pédagogique",
        "Citation sources",
        "Score global",
    ]
    cols_criteres = ["G","H","I","J","K","L"]

    for i, (nom, col) in enumerate(zip(criteres_noms, cols_criteres), 3):
        n_fiches = len(fiches)
        bg = GREY if i % 2 == 0 else WHITE

        dcell(ws_synth, i, 1, nom, bg=bg, bold=True)

        # Référence aux onglets annotateurs
        c2 = ws_synth.cell(
            row=i, column=2,
            value=f"='Annotateur 1'!{col}{n_fiches+4}"
        )
        c2.font      = Font(size=11, name="Arial", bold=True)
        c2.fill      = PatternFill("solid", fgColor=bg)
        c2.alignment = Alignment(horizontal="center")
        c2.border    = border()
        c2.number_format = "0.00"

        c3 = ws_synth.cell(
            row=i, column=3,
            value=f"='Annotateur 2'!{col}{n_fiches+4}"
        )
        c3.font      = Font(size=11, name="Arial", bold=True)
        c3.fill      = PatternFill("solid", fgColor=bg)
        c3.alignment = Alignment(horizontal="center")
        c3.border    = border()
        c3.number_format = "0.00"

        c4 = ws_synth.cell(
            row=i, column=4,
            value=f"=IFERROR(ABS(B{i}-C{i}),\"\")"
        )
        c4.font      = Font(size=11, name="Arial", bold=True)
        c4.fill      = PatternFill("solid", fgColor=bg)
        c4.alignment = Alignment(horizontal="center")
        c4.border    = border()
        c4.number_format = "0.00"

        ws_synth.row_dimensions[i].height = 22

    # Note κ Cohen
    r_note = len(criteres_noms) + 4
    ws_synth.merge_cells(f"A{r_note}:D{r_note}")
    c = ws_synth.cell(row=r_note, column=1,
                      value="→ Calculer le κ Cohen : python evaluation/compute_kappa.py")
    c.font      = Font(italic=True, size=10, color="444444", name="Arial")
    c.fill      = PatternFill("solid", fgColor=LIGHT)
    c.alignment = Alignment(horizontal="center")

    wb.save(OUTPUT)
    print(f"\n✓ Formulaire Excel généré : {OUTPUT}")
    print(f"  {len(fiches)} fiches | 2 onglets annotateurs | 1 synthèse")
    print(f"\n  Instructions :")
    print(f"  1. Ouvrir le fichier Excel")
    print(f"  2. Onglet 'Guide de notation' — lire la grille")
    print(f"  3. Onglet 'Annotateur 1' — noter les 5 premières fiches en calibration")
    print(f"  4. Onglet 'Annotateur 2' — l'encadreur note de son côté")
    print(f"  5. Onglet 'Synthèse' — scores automatiques après notation")


if __name__ == "__main__":
    main()